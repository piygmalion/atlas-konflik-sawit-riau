#!/usr/bin/env python3
"""Audit coverage: workspace sources vs website/data (repo)."""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

ROOT = Path(r"C:\Users\Patron\Documents\Konflik Sawit di Riau")
WEB = ROOT / "website" / "data"


def count_csv(path: Path) -> int:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for _ in csv.reader(f)) - 1)


def sheet_info(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info = {s: wb[s].max_row or 0 for s in wb.sheetnames}
    wb.close()
    return info


def load_web():
    out = {}
    for p in sorted(WEB.glob("*")):
        if p.suffix not in {".json", ".geojson"}:
            continue
        d = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(d, dict):
            if "records" in d:
                n = len(d["records"])
            elif "features" in d:
                n = len(d["features"])
            elif "gfw_match_bps" in d:
                n = {
                    "gfw_match_bps": len(d["gfw_match_bps"].get("records", [])),
                    "atlas_match": len(d["atlas_match"].get("records", [])),
                    "kepmenhut_36_2025": len(d["kepmenhut_36_2025"].get("records", [])),
                }
            elif "counts" in d:
                n = d["counts"]
            else:
                n = list(d.keys())
        else:
            n = len(d)
        out[p.name] = {"bytes": p.stat().st_size, "payload": n}
    return out


def main():
    web = load_web()
    report = {
        "web_data": web,
        "included": [],
        "partial": [],
        "missing": [],
        "workspace_xlsx": {},
        "workspace_csv": {},
        "workspace_geojson": {},
    }

    # Workbooks
    for p in sorted(ROOT.glob("*.xlsx")):
        report["workspace_xlsx"][p.name] = sheet_info(p)

    for p in sorted(ROOT.glob("*.csv")):
        report["workspace_csv"][p.name] = count_csv(p)

    for p in list(ROOT.glob("*.geojson")) + list((ROOT / "tmp").rglob("*.geojson")):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            report["workspace_geojson"][str(p.relative_to(ROOT))] = len(d.get("features", []))
        except Exception as e:
            report["workspace_geojson"][str(p.relative_to(ROOT))] = f"error:{e}"

    # Mapping validation
    checks = [
        {
            "id": "kasus_konflik",
            "source": "TABEL_KONFLIK_AGRARIA_SAWIT_RIAU.xlsx / Master_Kasus_Sawit",
            "workspace_n": report["workspace_xlsx"].get("TABEL_KONFLIK_AGRARIA_SAWIT_RIAU.xlsx", {}).get("Master_Kasus_Sawit"),
            "web_n": web.get("kasus.json", {}).get("payload"),
            "status_hint": "included",
        },
        {
            "id": "objek_agrinas",
            "source": "master_list_objek_agrinas_satgas_riau.csv",
            "workspace_n": report["workspace_csv"].get("master_list_objek_agrinas_satgas_riau.csv"),
            "web_n": web.get("objek_agrinas.json", {}).get("payload"),
            "status_hint": "included",
        },
        {
            "id": "polres_ranking",
            "source": "ranking_potensi_konflik_per_polres.csv",
            "workspace_n": report["workspace_csv"].get("ranking_potensi_konflik_per_polres.csv"),
            "web_n": web.get("polres.json", {}).get("payload"),
            "status_hint": "included",
        },
        {
            "id": "kab_kota_cluster",
            "source": "cluster_kabkota_agrinas.csv",
            "workspace_n": report["workspace_csv"].get("cluster_kabkota_agrinas.csv"),
            "web_n": len(json.loads((WEB / "kab_kota.json").read_text(encoding="utf-8"))["records"]),
            "status_hint": "included",
        },
        {
            "id": "perusahaan_gabungan",
            "source": "daftar_perusahaan_sawit_riau_gabungan.csv",
            "workspace_n": report["workspace_csv"].get("daftar_perusahaan_sawit_riau_gabungan.csv"),
            "web_n": web.get("perusahaan.json", {}).get("payload"),
            "status_hint": "included",
        },
        {
            "id": "gfw_match_bps",
            "source": "tabulasi_konsesi_sawit_gfw_match_bps_riau.csv",
            "workspace_n": report["workspace_csv"].get("tabulasi_konsesi_sawit_gfw_match_bps_riau.csv"),
            "web_n": web.get("konsesi.json", {}).get("payload", {}).get("gfw_match_bps"),
            "status_hint": "included",
        },
        {
            "id": "atlas_match",
            "source": "cocokan_atlas_gabungan_gfw.csv",
            "workspace_n": report["workspace_csv"].get("cocokan_atlas_gabungan_gfw.csv"),
            "web_n": web.get("konsesi.json", {}).get("payload", {}).get("atlas_match"),
            "status_hint": "included",
        },
        {
            "id": "kepmenhut_36",
            "source": "tabulasi_konsesi_sawit_kepmenhut_36_2025_riau_rapi.csv",
            "workspace_n": report["workspace_csv"].get("tabulasi_konsesi_sawit_kepmenhut_36_2025_riau_rapi.csv"),
            "web_n": web.get("konsesi.json", {}).get("payload", {}).get("kepmenhut_36_2025"),
            "status_hint": "included",
        },
        {
            "id": "titik_agrinas_geo",
            "source": "proksi_peta_titik_agrinas.geojson",
            "workspace_n": report["workspace_geojson"].get("proksi_peta_titik_agrinas.geojson"),
            "web_n": "subset of layers.geojson (objek_titik)",
            "status_hint": "partial",
        },
        {
            "id": "koridor",
            "source": "proksi_peta_koridor_agrinas.csv",
            "workspace_n": report["workspace_csv"].get("proksi_peta_koridor_agrinas.csv"),
            "web_n": "subset of layers.geojson (koridor)",
            "status_hint": "partial",
        },
        {
            "id": "risiko_kab_from_tabel",
            "source": "TABEL_KONFLIK.../Peta_Risiko_Sawit_Kab",
            "workspace_n": report["workspace_xlsx"].get("TABEL_KONFLIK_AGRARIA_SAWIT_RIAU.xlsx", {}).get("Peta_Risiko_Sawit_Kab"),
            "web_n": "merged into kab_kota.json.risiko_register",
            "status_hint": "partial",
        },
    ]

    missing_sources = [
        ("Inventarisasi_Sawit_Riau_v1.1_20260730.xlsx", "data_kasus_utama / dashboard / referensi — tidak diekspor terpisah (kasus diambil dari TABEL_KONFLIK)"),
        ("Template_Inventarisasi_Permasalahan_Sawit_Riau.xlsx", "template kosong / operasional, bukan dataset publik"),
        ("Normalisasi_Perusahaan_Sawit_Riau.xlsx / v2", "profil normalisasi — digantikan daftar gabungan CSV"),
        ("Daftar_Perusahaan_Sawit_Riau.xlsx", "multi-sheet izin 2017 — hanya ringkas via perusahaan.json"),
        ("REKAPITULASI PERIZINAN PERKEBUNAN 2017 ok.xlsx", "izin per kab — belum masuk web"),
        ("Tabulasi_Penertiban_Kawasan_Hutan_Sawit_Riau.xlsx", "15 sheet PKH/Satgas/Tesso Nilo/denda — belum masuk web"),
        ("Baseline_Publik_Rantai_Satgas_Agrinas.xlsx", "rantai penyerahan — belum masuk web"),
        ("Master_List_Fase2_Agrinas_Satgas_Riau.xlsx", "fase 2 / gap matching — belum masuk web"),
        ("cluster_kabkota_proksi_peta_agrinas.xlsx", "metodologi + densitas KLHK — sebagian via CSV cluster"),
        ("tabulasi_perkiraan_lokasi_sebaran_riau.xlsx", "hotspot detail sebaran — belum masuk web"),
        ("tabulasi_verifikasi_lanjutan_sebaran_riau.xlsx", "GCP/verifikasi georef — belum masuk web"),
        ("Ranking_Potensi_Konflik_Per_Polres.xlsx", "ada CSV twin — sudah masuk via CSV"),
        ("Rencana_Perbaikan_Inventarisasi_Sawit_Riau.xlsx", "manajemen proyek internal — tidak untuk publik"),
        ("tabulasi_konsesi_sawit_gfw_bbox_riau.csv", "287 konsesi bbox — HANYA match BPS (126) yang diekspor"),
        ("tmp/spatial/gfw_oilpalm_riau.geojson", "poligon konsesi GFW — belum masuk web"),
        ("tmp/spatial/idn_adm2_simplified.geojson", "batas adm2 — belum masuk web"),
        ("tmp/spatial/red_clusters_georef.csv / verifikasi_per_kabupaten.csv", "verifikasi spasial lanjutan — belum masuk web"),
        ("tmp/spatial/desa_lock/*", "kunci desa sentinel — belum masuk web"),
    ]

    for c in checks:
        wn, we = c["workspace_n"], c["web_n"]
        ok = False
        if isinstance(wn, int) and isinstance(we, int):
            ok = wn == we
            c["match"] = ok
            c["delta"] = we - wn
        else:
            c["match"] = None
        if c["status_hint"] == "included" and ok:
            report["included"].append(c)
        elif c["status_hint"] == "partial" or (isinstance(wn, int) and isinstance(we, int) and not ok):
            report["partial"].append(c)
        else:
            report["partial"].append(c)

    report["missing"] = [{"source": a, "note": b} for a, b in missing_sources]

    # Coverage score rough
    core_ok = sum(1 for c in report["included"] if c.get("match") is True)
    core_total = 8
    report["summary"] = {
        "core_datasets_matched": f"{core_ok}/{core_total}",
        "web_files": list(web.keys()),
        "xlsx_count": len(report["workspace_xlsx"]),
        "csv_count": len(report["workspace_csv"]),
        "verdict": (
            "Repo memuat INTI analitik (kasus, objek, polres, kab, perusahaan, konsesi match, lapisan proksi), "
            "tetapi BUKAN seluruh workbook/CSV/geojson workspace. Celah utama: penertiban KH, izin 2017 per kab, "
            "GFW bbox penuh + poligon, verifikasi sebaran/GCP, fase2 Agrinas, baseline rantai Satgas, batas admin."
        ),
    }

    out = ROOT / "website" / "_audit_coverage.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    print("\nINCLUDED:")
    for c in report["included"]:
        print(f"  OK {c['id']}: ws={c['workspace_n']} web={c['web_n']}")
    print("\nPARTIAL:")
    for c in report["partial"]:
        print(f"  ~ {c['id']}: ws={c['workspace_n']} web={c['web_n']}")
    print("\nMISSING (count):", len(report["missing"]))
    for m in report["missing"]:
        print(f"  - {m['source']}: {m['note']}")
    print("\nwrote", out)


if __name__ == "__main__":
    main()
