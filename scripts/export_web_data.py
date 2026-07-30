#!/usr/bin/env python3
"""
Ekspor workbook/CSV workspace -> website/data/*.json|geojson
Jalankan berkala setelah update workbook:

  python website/scripts/export_web_data.py
"""

from __future__ import annotations

import csv
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl diperlukan: pip install openpyxl") from exc

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    from shapely.geometry import mapping, shape
    from shapely.ops import transform as shp_transform

    HAS_SHAPELY = True
except ImportError:
    HAS_SHAPELY = False

HERE = Path(__file__).resolve()
SITE = HERE.parents[1]
_candidates = [SITE.parent, SITE]
ROOT = next(
    (
        p
        for p in _candidates
        if (p / "Master_List_Objek_Agrinas_Satgas_Riau.xlsx").exists()
        or (p / "master_list_objek_agrinas_satgas_riau.csv").exists()
    ),
    SITE.parent,
)
OUT = SITE / "data"
OUT.mkdir(parents=True, exist_ok=True)

RIAU_BBOX = (100.0, -1.2, 103.6, 2.6)
RIAU_ADM2_RAW = [
    "bengkalis",
    "dumai",
    "kota dumai",
    "indragiri hilir",
    "indragiri hulu",
    "kampar",
    "kepulauan meranti",
    "kuantan singingi",
    "pelalawan",
    "pekanbaru",
    "kota pekanbaru",
    "rokan hilir",
    "rokan hulu",
    "siak",
]


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    return s if s and s.lower() not in {"none", "null", "-"} else None


def read_csv(name: str) -> list[dict]:
    path = ROOT / name
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        return [{k: clean(v) for k, v in row.items()} for row in csv.DictReader(f)]


def sheet_rows(wb_name: str, sheet: str, header_row: int = 1) -> list[dict]:
    path = ROOT / wb_name
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < header_row:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_row - 1])]
    out = []
    for row in rows[header_row:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        out.append({headers[i]: clean(row[i] if i < len(row) else None) for i in range(len(headers))})
    return out


def write_json(name: str, payload, compact: bool = False):
    path = OUT / name
    if compact:
        path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    else:
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  wrote {path.name} ({path.stat().st_size:,} bytes)")


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", ".").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


def norm_name(s: str | None) -> str:
    if not s:
        return ""
    t = str(s).lower()
    t = t.replace("kabupaten", " ").replace("kab.", " ").replace("kota", " ")
    t = t.replace("kepulauan", "kep").replace("kep.", "kep")
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


RIAU_ADM2_NAMES = {norm_name(x) for x in RIAU_ADM2_RAW}


def slug(s: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "unknown").lower()).strip("-")


def parse_bbox(text: str | None):
    if not text:
        return None
    t = str(text).lower().replace("–", "-").replace("—", "-")
    m = re.search(
        r"lon\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*;\s*lat\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)",
        t,
    )
    if m:
        xs = sorted([float(m.group(1)), float(m.group(2))])
        ys = sorted([float(m.group(3)), float(m.group(4))])
        return xs[0], ys[0], xs[1], ys[1]
    nums = [to_float(x) for x in re.split(r"[,;\s]+", t) if to_float(x) is not None]
    if len(nums) >= 4:
        xs = sorted([nums[0], nums[2]])
        ys = sorted([nums[1], nums[3]])
        return xs[0], ys[0], xs[1], ys[1]
    return None


def geom_centroid_lonlat(geom: dict):
    if not geom:
        return None, None
    if HAS_SHAPELY:
        try:
            c = shape(geom).centroid
            return float(c.x), float(c.y)
        except Exception:
            pass
    coords = []

    def walk(node):
        if isinstance(node, (list, tuple)) and node and isinstance(node[0], (int, float)):
            coords.append((float(node[0]), float(node[1])))
        elif isinstance(node, (list, tuple)):
            for child in node:
                walk(child)

    walk(geom.get("coordinates"))
    if not coords:
        return None, None
    return sum(x for x, _ in coords) / len(coords), sum(y for _, y in coords) / len(coords)


def densitas_level_from_skor(skor) -> str:
    """Derive densitas label from skor_komposit (0–100), not risiko_register.level."""
    try:
        s = float(skor or 0)
    except (TypeError, ValueError):
        s = 0.0
    if s >= 70:
        return "TINGGI"
    if s >= 40:
        return "SEDANG"
    return "RENDAH"


def in_riau_bbox(lon: float | None, lat: float | None) -> bool:
    if lon is None or lat is None:
        return False
    minx, miny, maxx, maxy = RIAU_BBOX
    return minx - 0.3 <= lon <= maxx + 0.3 and miny - 0.3 <= lat <= maxy + 0.3


def simplify_geometry(geom: dict, tolerance: float = 0.004):
    if not geom:
        return geom
    if HAS_SHAPELY:
        try:
            g = shape(geom)
            if g.is_empty:
                return geom
            simple = g.simplify(tolerance, preserve_topology=True)
            if simple.is_empty:
                return geom
            return mapping(simple)
        except Exception:
            return geom
    return geom


def export_kab_kota():
    clusters = read_csv("cluster_kabkota_agrinas.csv")
    risiko = sheet_rows("TABEL_KONFLIK_AGRARIA_SAWIT_RIAU.xlsx", "Peta_Risiko_Sawit_Kab")
    risiko_map = {}
    for r in risiko:
        kab = r.get("Kabupaten_Kota")
        if kab is not None and str(kab).strip():
            risiko_map[str(kab).lower()] = r

    records = []
    for c in clusters:
        kab = c.get("kab_kota")
        lon = to_float(c.get("lon_centroid_proksi"))
        lat = to_float(c.get("lat_centroid_proksi"))
        risk = {}
        for key, val in risiko_map.items():
            if kab and (kab.lower() in key or key in kab.lower()):
                risk = val
                break
        records.append(
            {
                "id": slug(kab),
                "kab_kota": kab,
                "shape_name": None,
                "cluster": c.get("cluster"),
                "skor_komposit": to_float(c.get("skor_komposit")),
                "kategori_peta": c.get("kategori_peta"),
                "sinyal_agrinas": to_float(c.get("sinyal_agrinas_0_5")),
                "sebaran": to_float(c.get("sebaran_0_5")),
                "kasus_konflik_agrinas": to_float(c.get("kasus_konflik_agrinas")),
                "luas_disebut_terbuka": c.get("luas_disebut_terbuka"),
                "klhk_korp_kh_2022_ha": to_float(c.get("klhk_korp_kh_2022_ha")),
                "objek_sinyal_utama": c.get("objek_sinyal_utama"),
                "hotspot_kecamatan": c.get("hotspot_kecamatan_perkiraan"),
                "polres_proksi": c.get("polres_proksi"),
                "ketidakpastian": c.get("ketidakpastian"),
                "lon": lon,
                "lat": lat,
                "catatan_peta": c.get("catatan_peta"),
                "n_kasus": 0,
                "risiko_register": {
                    "skor": to_float(risk.get("Skor_Risiko_Sawit")),
                    "level": risk.get("Level_Risiko"),
                    "kasus_ops": to_float(risk.get("Jumlah_Kasus_Ops_Sawit")),
                    "jumlah_lp": to_float(risk.get("Jumlah_LP")),
                    "driver_utama": risk.get("Driver_Utama"),
                    "rekomendasi": risk.get("Rekomendasi"),
                },
            }
        )
    return records


def match_kab_record(records: list[dict], name: str):
    n = norm_name(name)
    for rec in records:
        rn = norm_name(rec.get("kab_kota"))
        if not rn:
            continue
        if n == rn or n in rn or rn in n:
            return rec
    return None


def export_adm2_choropleth(kab_records: list[dict]):
    src = ROOT / "tmp" / "spatial" / "idn_adm2_simplified.geojson"
    features = []
    if not src.exists():
        write_json("adm2_riau.geojson", {"type": "FeatureCollection", "features": []}, compact=True)
        return []

    raw = json.loads(src.read_text(encoding="utf-8"))
    used = set()
    for f in raw.get("features", []):
        shape_name = (f.get("properties") or {}).get("shapeName")
        n = norm_name(shape_name)
        if n not in RIAU_ADM2_NAMES and not any(n == x or x in n or n in x for x in RIAU_ADM2_NAMES):
            continue
        lon, lat = geom_centroid_lonlat(f.get("geometry"))
        if not in_riau_bbox(lon, lat):
            continue
        rec = match_kab_record(kab_records, shape_name)
        if not rec:
            continue
        used.add(rec["id"])
        rec["shape_name"] = shape_name
        if lon and lat and (not rec.get("lon") or not rec.get("lat")):
            rec["lon"], rec["lat"] = lon, lat
        skor = rec.get("skor_komposit") or 0
        risk_skor = (rec.get("risiko_register") or {}).get("skor")
        features.append(
            {
                "type": "Feature",
                "geometry": f.get("geometry"),
                "properties": {
                    "id": rec["id"],
                    "nama": rec["kab_kota"],
                    "shape_name": shape_name,
                    "skor": skor,
                    "skor_risiko": risk_skor,
                    "level_risiko": (rec.get("risiko_register") or {}).get("level"),
                    "kategori": rec.get("kategori_peta"),
                    "polres": rec.get("polres_proksi"),
                    "n_kasus": rec.get("n_kasus") or 0,
                    "layer": "choropleth",
                },
            }
        )

    write_json("adm2_riau.geojson", {"type": "FeatureCollection", "features": features}, compact=True)
    print(f"    choropleth polygons: {len(features)} (matched kab ids: {len(used)})")
    return features


def export_polres():
    rows = read_csv("ranking_potensi_konflik_per_polres.csv")
    records = []
    for r in rows:
        records.append(
            {
                "peringkat": int(to_float(r.get("peringkat")) or 0),
                "polres": r.get("polres"),
                "skor": to_float(r.get("skor")),
                "kategori": r.get("kategori"),
                "skor_osint": to_float(r.get("skor_osint_saja")),
                "skor_register": to_float(r.get("skor_register_saja")),
                "n_entri": to_float(r.get("n_entri_terpetakan")),
                "n_aksi_massa": to_float(r.get("n_aksi_massa")),
                "n_kekerasan": to_float(r.get("n_kekerasan")),
                "n_agrinas": to_float(r.get("n_agrinas_satgas_kso")),
                "n_recent": to_float(r.get("n_recent_2024plus")),
                "n_tabel_konflik": to_float(r.get("n_tabel_konflik_agraria")),
                "komponen": {
                    "liputan": to_float(r.get("komponen_liputan")),
                    "aksi": to_float(r.get("komponen_aksi")),
                    "objek": to_float(r.get("komponen_objek")),
                    "status": to_float(r.get("komponen_status")),
                    "adat": to_float(r.get("komponen_adat")),
                },
                "alasan": r.get("alasan_singkat"),
                "tahun": r.get("tahun_tercover"),
            }
        )

    coverage = {
        "total_entri_terpetakan": None,
        "entri_tidak_terpetakan": None,
        "bucket_tidak_terpetakan": "Lintas Provinsi Riau / tidak terpetakan Polres",
        "label_terpetakan": "Entri terpetakan ke Polres",
        "label_tidak_terpetakan": "Lintas Provinsi Riau / tidak terpetakan Polres",
    }
    model_catatan = (
        "Skor adalah indeks early-warning dari liputan OSINT + objek Agrinas/KSO + register terbuka, "
        "bukan vonis operasional. Belum dikalibrasi ulang terhadap rekap LP/SPKT 36 bulan resmi."
    )
    ranking_json = ROOT / "ranking_potensi_konflik_per_polres.json"
    if ranking_json.exists():
        try:
            raw = json.loads(ranking_json.read_text(encoding="utf-8"))
            coverage.update(raw.get("ringkasan") or {})
            if raw.get("model", {}).get("catatan"):
                model_catatan = (
                    "Indeks liputan+objek+register terbuka — bukan vonis operasional. "
                    + str(raw["model"]["catatan"])
                )
        except Exception:
            pass

    write_json(
        "polres.json",
        {
            "model": {
                "kategori": {"PANTAU": "0-39", "WASPADA": "40-69", "PRIORITAS": "70-100"},
                "blend_default": {"osint": 0.7, "register": 0.3},
                "catatan": model_catatan,
            },
            "coverage": coverage,
            "records": records,
        },
    )
    return records


def patch_polres_coverage(kasus: list[dict], polres_records: list[dict]):
    """Recompute coverage from current kasus serving set (post-DQ)."""
    lintas = sum(
        1
        for r in kasus
        if "lintas" in str(r.get("polres") or "").lower()
        or "lintas" in str(r.get("kab_kota") or "").lower()
    )
    mapped = len(kasus) - lintas
    path = OUT / "polres.json"
    payload = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {"records": polres_records}
    cov = payload.get("coverage") or {}
    cov.update(
        {
            "n_polres": len(payload.get("records") or polres_records),
            "prioritas": sum(1 for p in (payload.get("records") or []) if p.get("kategori") == "PRIORITAS"),
            "waspada": sum(1 for p in (payload.get("records") or []) if p.get("kategori") == "WASPADA"),
            "pantau": sum(1 for p in (payload.get("records") or []) if p.get("kategori") == "PANTAU"),
            "total_entri_terpetakan": mapped,
            "entri_tidak_terpetakan": lintas,
            "bucket_tidak_terpetakan": "Lintas Provinsi Riau / tidak terpetakan Polres",
            "label_terpetakan": "Entri terpetakan ke Polres",
            "label_tidak_terpetakan": "Lintas Provinsi Riau / tidak terpetakan Polres",
        }
    )
    payload["coverage"] = cov
    write_json("polres.json", payload)
    return cov


def export_objek():
    rows = read_csv("master_list_objek_agrinas_satgas_riau.csv")
    records = [
        {
            "id": r.get("id"),
            "nama": r.get("nama_kanonik"),
            "tipe_badan": r.get("tipe_badan"),
            "lapisan": r.get("lapisan"),
            "lapisan_semua": r.get("lapisan_semua"),
            "peran": r.get("peran"),
            "klaster": r.get("klaster"),
            "kab_kota": r.get("kab_kota"),
            "kab_primary": r.get("kab_primary"),
            "kab_list": r.get("kab_list"),
            "polres_primary": r.get("polres_primary"),
            "mappable": r.get("mappable"),
            "luas_disebut": r.get("luas_disebut"),
            "status_kredibilitas": r.get("status_kredibilitas"),
            "prioritas": r.get("prioritas"),
            "kaitan_agrinas": r.get("kaitan_agrinas"),
            "cro_regional": r.get("cro_regional"),
            "mitra_pair": r.get("mitra_pair"),
            "ada_di_bps": r.get("ada_di_bps"),
            "ada_di_konflik_polda": r.get("ada_di_konflik_polda"),
            "sumber": r.get("sumber"),
        }
        for r in rows
    ]
    write_json("objek_agrinas.json", {"total": len(records), "records": records})
    return records


def export_kasus():
    """Prefer DQ-cleaned CSV; fall back to workbook sheet."""
    records = []
    csv_rows = read_csv("master_kasus_sawit_riau.csv")
    if csv_rows:
        for r in csv_rows:
            if not r.get("id"):
                continue
            # skip noise if reintroduced
            if str(r.get("status_verifikasi") or "").lower() == "noise":
                continue
            uraian = str(r.get("uraian") or "")
            if re.search(r"historikonflik\s*nihil|nihil\s*nihil", uraian, re.I):
                continue
            rec = {
                "id": r.get("id"),
                "sumber_dokumen": r.get("sumber_dokumen"),
                "polres": r.get("polres"),
                "kab_kota": r.get("kab_kota"),
                "tahun": r.get("tahun"),
                "nomor_lp": r.get("nomor_lp"),
                "jenis": r.get("jenis"),
                "kategori": r.get("kategori"),
                "pihak": r.get("pihak"),
                "lokasi": r.get("lokasi"),
                "uraian": r.get("uraian"),
                "upaya": r.get("upaya"),
                "status": r.get("status"),
                "hambatan": r.get("hambatan"),
                "perusahaan": r.get("perusahaan"),
                "tema": r.get("tema"),
                "tipe_entri": r.get("tipe_entri"),
                "indikator_sawit": r.get("indikator_sawit"),
                "tanpa_lp": str(r.get("tanpa_lp") or "").lower() in {"true", "1", "ya"},
                "tanpa_lp_alasan": r.get("tanpa_lp_alasan"),
                "status_verifikasi": r.get("status_verifikasi") or "ok",
            }
            records.append(rec)
    else:
        rows = sheet_rows("TABEL_KONFLIK_AGRARIA_SAWIT_RIAU.xlsx", "Master_Kasus_Sawit")
        for r in rows:
            if not r.get("ID") and not r.get("Uraian_Singkat"):
                continue
            if not r.get("ID"):
                continue
            records.append(
                {
                    "id": r.get("ID"),
                    "sumber_dokumen": r.get("Sumber_Dokumen"),
                    "polres": r.get("Wilayah_Polres"),
                    "kab_kota": r.get("Kabupaten_Kota"),
                    "tahun": r.get("Tahun_Referensi"),
                    "nomor_lp": r.get("Nomor_LP"),
                    "jenis": r.get("Jenis_Konflik"),
                    "kategori": r.get("Kategori_Konflik"),
                    "pihak": r.get("Pihak_Berkonflik"),
                    "lokasi": r.get("Lokasi"),
                    "uraian": r.get("Uraian_Singkat"),
                    "upaya": r.get("Upaya_Dilakukan"),
                    "status": r.get("Status_Keterangan"),
                    "hambatan": r.get("Hambatan"),
                    "perusahaan": r.get("Perusahaan_Terkait"),
                    "tema": r.get("Kata_Kunci_Tema"),
                    "tipe_entri": r.get("Tipe_Entri"),
                    "indikator_sawit": r.get("Indikator_Sawit"),
                }
            )
    for rec in records:
        disebut = years_from_kasus(rec.get("tahun")) or ["2026"]
        kejadian = years_kejadian_kasus(rec) or disebut[:1]
        rec["tahun_disebut"] = disebut
        rec["tahun_kejadian"] = kejadian
        # light DQ normalize
        for k in ("nomor_lp", "status", "upaya", "uraian"):
            if rec.get(k) is not None:
                rec[k] = re.sub(r"[ \t]+", " ", str(rec[k]).replace("\r", "")).strip() or None
    write_json("kasus.json", {"total": len(records), "records": records})
    return records


def attach_kasus_counts(kab_records: list[dict], kasus: list[dict]):
    for rec in kab_records:
        n = 0
        for k in kasus:
            if match_wilayah_py(k.get("kab_kota"), rec.get("kab_kota")) or match_wilayah_py(
                k.get("polres"), rec.get("polres_proksi")
            ):
                n += 1
        rec["n_kasus"] = n


def match_wilayah_py(a, b) -> bool:
    if not a or not b:
        return False
    na, nb = norm_name(a), norm_name(b)
    aliases = {
        "rokan hulu": ["rohul"],
        "rokan hilir": ["rohil"],
        "indragiri hulu": ["inhu"],
        "indragiri hilir": ["inhil"],
        "kuantan singingi": ["kuansing"],
        "kepulauan meranti": ["meranti", "kep meranti"],
    }
    bag_a = {na}
    bag_b = {nb}
    for canon, als in aliases.items():
        if na == canon or any(x in na for x in als) or canon in na:
            bag_a |= {canon, *als}
        if nb == canon or any(x in nb for x in als) or canon in nb:
            bag_b |= {canon, *als}
    return any(x in y or y in x for x in bag_a for y in bag_b if x and y)


def export_spatial_layers(kab_records: list[dict]):
    features = []

    # Titik objek Agrinas — skip centroid REF (noise visual / ID duplikat historis)
    path = ROOT / "proksi_peta_titik_agrinas.geojson"
    if path.exists():
        raw = json.loads(path.read_text(encoding="utf-8"))
        for f in raw.get("features", []):
            props = {k: clean(v) for k, v in (f.get("properties") or {}).items()}
            prioritas = str(props.get("prioritas") or "").upper()
            tipe = str(props.get("tipe") or "").lower()
            nama = str(props.get("nama") or "").lower()
            if "REF" in prioritas or "centroid" in tipe or "centroid" in nama:
                continue
            features.append(
                {
                    "type": "Feature",
                    "geometry": f.get("geometry"),
                    "properties": {
                        "id": props.get("id"),
                        "nama": props.get("nama"),
                        "kab_kota": props.get("kab_kota"),
                        "tipe": props.get("tipe") or "objek",
                        "prioritas": props.get("prioritas"),
                        "catatan": props.get("catatan"),
                        "polres_proksi": props.get("polres_proksi"),
                        "sumber": props.get("sumber"),
                        "layer": "objek_titik",
                    },
                }
            )

    # Koridor dari bbox teks; fallback envelope dari anggota kab
    koridor_n = 0
    for k in read_csv("proksi_peta_koridor_agrinas.csv"):
        bbox = parse_bbox(k.get("bbox_approx"))
        if not bbox:
            members = re.split(r"[;,]", k.get("anggota_kab") or "")
            pts = []
            for m in members:
                rec = match_kab_record(kab_records, m.strip())
                if rec and rec.get("lon") is not None and rec.get("lat") is not None:
                    pts.append((rec["lon"], rec["lat"]))
            if len(pts) >= 1:
                xs = [p[0] for p in pts]
                ys = [p[1] for p in pts]
                pad = 0.25 if len(pts) == 1 else 0.15
                bbox = (min(xs) - pad, min(ys) - pad, max(xs) + pad, max(ys) + pad)
        if not bbox:
            continue
        minx, miny, maxx, maxy = bbox
        koridor_n += 1
        features.append(
            {
                "type": "Feature",
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [
                        [
                            [minx, miny],
                            [maxx, miny],
                            [maxx, maxy],
                            [minx, maxy],
                            [minx, miny],
                        ]
                    ],
                },
                "properties": {
                    "id": k.get("id"),
                    "nama": k.get("nama"),
                    "tipe": "koridor_bbox_proksi",
                    "anggota_kab": k.get("anggota_kab"),
                    "polres_proksi": k.get("polres_proksi"),
                    "karakter": k.get("karakter"),
                    "prioritas": k.get("prioritas_peta"),
                    "catatan": "Bounding box analitis — bukan koridor geografis resmi.",
                    "layer": "koridor",
                },
            }
        )

    # Densitas kasus (proksi ke centroid kab)
    densitas_n = 0
    for rec in kab_records:
        if rec.get("lon") is None or rec.get("lat") is None:
            continue
        n = int(rec.get("n_kasus") or 0)
        if n <= 0:
            continue
        densitas_n += 1
        features.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [rec["lon"], rec["lat"]]},
                "properties": {
                    "id": rec["id"],
                    "nama": rec["kab_kota"],
                    "n_kasus": n,
                    "weight": n,
                    "skor": rec.get("skor_komposit"),
                    "level_risiko": densitas_level_from_skor(rec.get("skor_komposit")),
                    "skor_register": (rec.get("risiko_register") or {}).get("skor"),
                    "layer": "densitas_kasus",
                },
            }
        )

    geo = {"type": "FeatureCollection", "features": features}
    write_json("layers.geojson", geo, compact=True)
    print(f"    layers: objek={sum(1 for f in features if f['properties']['layer']=='objek_titik')} "
          f"koridor={koridor_n} densitas={densitas_n}")
    return geo


def export_gfw_overlay():
    """Export light TopoJSON overlay for GitHub Pages (mapshaper), with GeoJSON fallback."""
    import shutil
    import subprocess

    src = ROOT / "tmp" / "spatial" / "gfw_oilpalm_riau.geojson"
    out_topo = OUT / "gfw_konsesi.topojson"
    out_geo = OUT / "gfw_konsesi.geojson"
    features = []

    if not src.exists():
        write_json(
            "gfw_konsesi.topojson",
            {"type": "Topology", "objects": {"gfw_konsesi": {"type": "GeometryCollection", "geometries": []}}, "arcs": []},
            compact=True,
        )
        if out_geo.exists():
            out_geo.unlink()
        print("    gfw polygons: 0 (source missing)")
        return features

    mapshaper = shutil.which("mapshaper") or shutil.which("npx")
    if mapshaper:
        cmd = [
            mapshaper,
            *(["-y", "mapshaper"] if Path(mapshaper).name.lower().startswith("npx") else []),
            str(src),
            "-filter-fields",
            "name,company,group_comp,area_ha,type,po_hgu",
            "-rename-fields",
            "group=group_comp,hgu=po_hgu",
            "-each",
            "layer='gfw_konsesi'",
            "-simplify",
            "weighted",
            "8%",
            "keep-shapes",
            "-rename-layers",
            "gfw_konsesi",
            "-o",
            "force",
            "format=topojson",
            "quantization=5000",
            str(out_topo),
        ]
        try:
            subprocess.run(cmd, check=True, capture_output=True, text=True)
            topo = json.loads(out_topo.read_text(encoding="utf-8"))
            n = len((topo.get("objects") or {}).get("gfw_konsesi", {}).get("geometries") or [])
            if out_geo.exists():
                out_geo.unlink()
            kb = out_topo.stat().st_size / 1024
            print(f"    gfw TopoJSON: {n} polygons · {kb:.0f} KB (mapshaper)")
            return [{"_": i} for i in range(n)]  # count placeholder for meta
        except (subprocess.CalledProcessError, OSError, json.JSONDecodeError) as exc:
            print(f"    mapshaper failed ({exc}); falling back to GeoJSON simplify")

    raw = json.loads(src.read_text(encoding="utf-8"))
    for f in raw.get("features", []):
        props = f.get("properties") or {}
        geom = simplify_geometry(f.get("geometry"), tolerance=0.008)
        features.append(
            {
                "type": "Feature",
                "geometry": geom,
                "properties": {
                    "name": props.get("name") or props.get("company"),
                    "company": props.get("company"),
                    "group": props.get("group_comp"),
                    "area_ha": props.get("area_ha"),
                    "type": props.get("type"),
                    "hgu": props.get("po_hgu"),
                    "layer": "gfw_konsesi",
                },
            }
        )
    # Compact GeoJSON fallback when TopoJSON tooling unavailable
    write_json("gfw_konsesi.geojson", {"type": "FeatureCollection", "features": features}, compact=True)
    print(f"    gfw GeoJSON fallback: {len(features)} (shapely={HAS_SHAPELY})")
    return features


def bucket_jenis(jenis: str | None, kategori: str | None) -> str:
    t = f"{jenis or ''} {kategori or ''}".lower()
    if "situasional" in t or "agrinas" in t or "kso" in t:
        return "Situasional KSO/Agrinas"
    if "sengketa" in t or "lahan" in t or "okupasi" in t or "plasma" in t:
        return "Sengketa lahan"
    if "ekonomi" in t:
        return "Ekonomi"
    if "kekerasan" in t or "bentrok" in t or "demo" in t or "unjuk" in t:
        return "Aksi / kekerasan"
    return "Lainnya"


def years_from_kasus(tahun) -> list[str]:
    if tahun is None:
        return []
    found = re.findall(r"20\d{2}", str(tahun))
    return sorted({y for y in found if 2018 <= int(y) <= 2026})


def years_kejadian_kasus(rec: dict) -> list[str]:
    chunks = [rec.get("uraian"), rec.get("nomor_lp"), rec.get("status"), rec.get("lokasi")]
    found: list[str] = []
    for c in chunks:
        found.extend(years_from_kasus(c))
    lp = str(rec.get("nomor_lp") or "")
    found.extend(re.findall(r"/?(20\d{2})/", lp))
    uniq = sorted({y for y in found if 2018 <= int(y) <= 2026})
    if uniq:
        return uniq
    disebut = years_from_kasus(rec.get("tahun"))
    return [min(disebut)] if disebut else []


def export_analytics(polres: list[dict], objek: list[dict], kasus: list[dict]):
    from collections import Counter

    # 1) Polres komponen for small-multiples / grouped bars
    polres_komponen = [
        {
            "polres": p.get("polres"),
            "label": str(p.get("polres") or "").replace("Polres ", ""),
            "peringkat": p.get("peringkat"),
            "skor": p.get("skor"),
            "kategori": p.get("kategori"),
            "komponen": p.get("komponen") or {},
        }
        for p in polres
    ]

    # 2) Agrinas layer flow (nodes + links)
    layer_order = [
        "A. Pengelola",
        "B. Eks pengelola",
        "C. Mitra KSO",
        "D. Eks lahan (via KSO)",
        "E. Gelombang 1 Satgas",
        "F. Objek kawasan",
    ]
    layer_counts = {k: 0 for k in layer_order}
    for o in objek:
        lap = o.get("lapisan") or "Lainnya"
        if lap not in layer_counts:
            layer_counts[lap] = 0
        layer_counts[lap] += 1

    nodes = [{"id": k, "label": k, "value": layer_counts.get(k, 0)} for k in layer_order if layer_counts.get(k, 0)]
    links = []

    def add_link(src, tgt, value, note=""):
        if value > 0:
            links.append({"source": src, "target": tgt, "value": int(value), "note": note})

    a = layer_counts.get("A. Pengelola", 0)
    b = layer_counts.get("B. Eks pengelola", 0)
    c = layer_counts.get("C. Mitra KSO", 0)
    d = layer_counts.get("D. Eks lahan (via KSO)", 0)
    e = layer_counts.get("E. Gelombang 1 Satgas", 0)
    f = layer_counts.get("F. Objek kawasan", 0)
    add_link("A. Pengelola", "C. Mitra KSO", min(a * 8, c) or a, "pengelola → mitra operasional")
    add_link("B. Eks pengelola", "C. Mitra KSO", min(b * 2, max(c - a, 0)) or min(b, c), "eks pengelola → mitra/KSO")
    add_link("C. Mitra KSO", "D. Eks lahan (via KSO)", min(c, d), "mitra → eks lahan via KSO")
    add_link("A. Pengelola", "E. Gelombang 1 Satgas", min(a, e) or (1 if a and e else 0), "jalur satgas")
    add_link("C. Mitra KSO", "E. Gelombang 1 Satgas", max(e - a, 0), "mitra dalam gelombang 1")
    add_link("D. Eks lahan (via KSO)", "F. Objek kawasan", min(d, f) or (1 if f else 0), "eks lahan ↔ objek kawasan")

    try:
        konsesi = json.loads((OUT / "konsesi.json").read_text(encoding="utf-8"))
    except Exception:
        konsesi = {"atlas_match": {"records": []}, "kepmenhut_36_2025": {"records": []}}

    atlas_rows = []
    for r in konsesi.get("atlas_match", {}).get("records", []):
        status = "cocok" if "cocok" in str(r.get("status") or "").lower() else (r.get("status") or "lain")
        konflik = (
            "ada di konflik"
            if str(r.get("ada_di_konflik_polda") or "").lower() in {"ya", "true", "1"}
            else "tidak di konflik"
        )
        atlas_rows.append(
            {
                "atlas_nama": r.get("atlas_nama"),
                "nama_lokal": r.get("nama_lokal"),
                "status_match": status,
                "konflik": konflik,
                "tahun": r.get("tahun"),
                "area_ha": r.get("area_ha"),
            }
        )

    atlas_flow_links = []
    match_c = Counter(r["status_match"] for r in atlas_rows)
    for st, n in match_c.items():
        atlas_flow_links.append({"source": "Nusantara Atlas", "target": st, "value": n})
    conf_c = Counter((r["status_match"], r["konflik"]) for r in atlas_rows)
    for (st, conf), n in conf_c.items():
        atlas_flow_links.append({"source": st, "target": conf, "value": n})

    kepmen_by_status = {}
    kepmen_records = []
    for r in konsesi.get("kepmenhut_36_2025", {}).get("records", []):
        st = r.get("status") or "Tidak diketahui"
        kepmen_by_status[st] = kepmen_by_status.get(st, 0) + 1
        kepmen_records.append(r)

    kepmen_buckets = {"Berproses": 0, "Ditolak": 0, "Campuran": 0, "Lainnya": 0}
    for st, n in kepmen_by_status.items():
        s = st.lower()
        if "campuran" in s:
            kepmen_buckets["Campuran"] += n
        elif "ditolak" in s and "berproses" not in s:
            kepmen_buckets["Ditolak"] += n
        elif "berproses" in s:
            kepmen_buckets["Berproses"] += n
        else:
            kepmen_buckets["Lainnya"] += n

    series_keys = ["Situasional KSO/Agrinas", "Sengketa lahan", "Ekonomi", "Aksi / kekerasan", "Lainnya"]

    def build_tl(year_fn):
        tl = {}
        tl_pol = {}
        for k in kasus:
            ys = year_fn(k) or ["2026"]
            bucket = bucket_jenis(k.get("jenis"), k.get("kategori"))
            pol = str(k.get("polres") or "Lain/Polda").replace("Polres ", "")
            for y in ys:
                tl.setdefault(y, {kk: 0 for kk in series_keys})
                tl[y][bucket] = tl[y].get(bucket, 0) + 1
                tl_pol.setdefault(y, {})
                tl_pol[y][pol] = tl_pol[y].get(pol, 0) + 1
        return sorted(tl.keys()), tl, tl_pol

    y_d, tj_d, tp_d = build_tl(lambda k: k.get("tahun_disebut") or years_from_kasus(k.get("tahun")))
    y_k, tj_k, tp_k = build_tl(lambda k: k.get("tahun_kejadian") or years_kejadian_kasus(k))

    payload = {
        "polres_komponen": polres_komponen,
        "agrinas_flow": {"nodes": nodes, "links": links, "counts": layer_counts},
        "atlas_flow": {"links": atlas_flow_links, "records": atlas_rows},
        "timeline": {
            "default_mode": "kejadian",
            "years": y_k,
            "categories": series_keys,
            "by_jenis": tj_k,
            "by_polres": tp_k,
            "disebut": {"years": y_d, "by_jenis": tj_d, "by_polres": tp_d},
            "kejadian": {"years": y_k, "by_jenis": tj_k, "by_polres": tp_k},
            "catatan": (
                "Mode kejadian = tahun di uraian/LP; mode disebut = Tahun_Referensi. "
                "Spike bisa tetap bias pengumpulan."
            ),
        },
        "kepmenhut": {
            "buckets": [{"label": k, "value": v} for k, v in kepmen_buckets.items() if v],
            "by_status_raw": [
                {"label": k, "value": v} for k, v in sorted(kepmen_by_status.items(), key=lambda x: -x[1])
            ],
            "records": kepmen_records,
            "total": len(kepmen_records),
        },
    }
    write_json("analytics.json", payload)
    print(f"    analytics: polres={len(polres_komponen)} timeline_kejadian={y_k} disebut={y_d} kepmen={len(kepmen_records)}")
    return payload


def parse_numeric(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(",", "")
    # Indonesian thousand dots: 1.400.000
    if re.fullmatch(r"\d{1,3}(\.\d{3})+(,\d+)?", s):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


def extract_table_from_sheet(wb_name: str, sheet: str, min_header_cells: int = 3) -> list[dict]:
    """Find first row with enough non-empty cells as header, then read records."""
    path = ROOT / wb_name
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    header_idx = None
    headers = []
    for i, row in enumerate(rows[:25]):
        cells = [clean(v) for v in row]
        nonempty = [c for c in cells if c]
        if len(nonempty) >= min_header_cells and not all(str(c).startswith("http") for c in nonempty[:1]):
            # Prefer rows that look like headers (many strings, few pure numbers)
            numericish = sum(1 for c in nonempty if isinstance(c, (int, float)) or re.fullmatch(r"-?\d+(\.\d+)?", str(c) or ""))
            if numericish <= max(1, len(nonempty) // 3):
                header_idx = i
                headers = []
                for j, c in enumerate(cells):
                    headers.append(str(c).strip() if c else f"col_{j}")
                break
    if header_idx is None:
        return []
    out = []
    for row in rows[header_idx + 1 :]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {}
        empty = True
        for j, h in enumerate(headers):
            val = clean(row[j] if j < len(row) else None)
            if val is not None:
                empty = False
            item[h] = val
        if empty:
            continue
        # skip section titles that only fill first column
        vals = [v for v in item.values() if v is not None]
        if len(vals) == 1 and isinstance(vals[0], str) and len(vals[0]) > 60:
            continue
        out.append(item)
    return out


def export_penertiban():
    wb_name = "Tabulasi_Penertiban_Kawasan_Hutan_Sawit_Riau.xlsx"
    wb_path = ROOT / wb_name
    # If workbook absent, keep existing penertiban.json (DQ-fixed) but refresh sk36 from CSV
    if not wb_path.exists():
        existing = {}
        path = OUT / "penertiban.json"
        if path.exists():
            existing = json.loads(path.read_text(encoding="utf-8"))
        sk36 = []
        for r in read_csv("tabulasi_sk36_2025_110a_riau_dq.csv"):
            sk36.append(
                {
                    "no": int(to_float(r.get("no")) or 0),
                    "no_partition": to_float(r.get("no_partition")),
                    "nama": r.get("nama"),
                    "dimohon_ha": to_float(r.get("dimohon_ha")),
                    "berproses_ha": to_float(r.get("berproses_ha")),
                    "ditolak_ha": to_float(r.get("ditolak_ha")),
                    "rasio_ditolak": r.get("rasio_ditolak"),
                    "prioritas": r.get("prioritas"),
                    "status_proses": r.get("status_proses"),
                    "status_proses_all": r.get("status_proses_all"),
                    "record_id": r.get("record_id"),
                }
            )
        if sk36:
            existing.setdefault("normalized", {})["sk36_2025_110a"] = {
                "total": len(sk36),
                "records": sk36,
                "pk": "record_id",
                "note": "Composite identity via record_id after DQ fix",
            }
        existing["updated_note"] = "Preserved existing penertiban (workbook absent); SK36 from DQ CSV"
        write_json("penertiban.json", existing)
        print(f"    penertiban: workbook missing — preserved; sk36={len(sk36)}")
        return existing

    sheets = {
        "estimasi_sawit_kh": "02_Estimasi_Sawit_KH",
        "fungsi_kawasan_eof": "03_Fungsi_Kawasan_EoF",
        "capaian_satgas_pkh": "04_Capaian_Satgas_PKH",
        "operasi_tesso_nilo": "05_Operasi_Tesso_Nilo",
        "kerangka_hukum": "06_Kerangka_Hukum",
        "kronologi": "07_Kronologi",
        "sebaran_lokal_inventaris": "08_Sebaran_Lokal_Inventaris",
        "denda_pnbp_110a_b": "09_Denda_PNBP_110A_B",
        "sumber_data": "10_Sumber_Data",
        "sebaran_kab_korporasi": "11_Sebaran_Kab_Korporasi",
        "satgas_tahap_residual": "12_Satgas_Tahap_Residual",
        "gelombang1_27_pt": "13_27_PT_Gelombang1_PKH",
        "sk36_2025_110a_riau": "14_SK36_2025_110A_Riau",
        "denda_dampak_sosial": "15_Denda_Dan_Dampak_Sosial",
        "status_celahan": "00_Status_Celahan",
        "ringkasan": "01_Ringkasan",
    }
    sections = {}
    for key, sheet in sheets.items():
        records = extract_table_from_sheet(wb_name, sheet)
        sections[key] = {"sheet": sheet, "total": len(records), "records": records}

    # Normalized convenience extracts
    kab_kh = []
    for r in sections.get("sebaran_kab_korporasi", {}).get("records", []):
        kab = r.get("Kabupaten/Kota") or r.get("Kabupaten") or r.get("col_0")
        if not kab:
            continue
        kab_s = str(kab).strip()
        if not (kab_s.lower().startswith("kab") or kab_s.lower().startswith("kota")):
            continue
        kab_kh.append(
            {
                "kab_kota": kab_s,
                "luas_ha": parse_numeric(r.get("Luas korporasi di KH tanpa izin (ha)")),
                "peringkat": parse_numeric(r.get("Peringkat")),
                "porsi": r.get("Porsi vs total korporasi Riau"),
                "catatan": r.get("Catatan operasi"),
            }
        )

    gelombang1 = []
    for r in sections.get("gelombang1_27_pt", {}).get("records", []):
        nama = r.get("Perusahaan") or r.get("Nama") or r.get("col_1")
        no = parse_numeric(r.get("No") or r.get("col_0"))
        if not nama or not isinstance(nama, str):
            continue
        if no is None or no < 1 or no > 40:
            continue
        if not re.search(r"\bPT\b|Perusahaan|Koperasi|CV\b", nama, re.I) and not nama.upper().startswith("PT"):
            # still allow plain company-like names with kabupaten filled
            if not (r.get("Kabupaten") or r.get("col_2")):
                continue
        gelombang1.append(
            {
                "no": int(no),
                "perusahaan": nama,
                "kabupaten": r.get("Kabupaten") or r.get("col_2"),
                "afiliasi": r.get("Afiliasi / grup (indikasi)") or r.get("col_3"),
                "catatan": r.get("Catatan luas / status terbuka") or r.get("col_4"),
            }
        )

    sk36 = []
    # Prefer DQ-cleaned CSV if present
    sk36_csv = read_csv("tabulasi_sk36_2025_110a_riau_dq.csv")
    if sk36_csv:
        for r in sk36_csv:
            sk36.append(
                {
                    "no": int(to_float(r.get("no")) or 0),
                    "no_partition": to_float(r.get("no_partition")),
                    "nama": r.get("nama"),
                    "dimohon_ha": to_float(r.get("dimohon_ha")),
                    "berproses_ha": to_float(r.get("berproses_ha")),
                    "ditolak_ha": to_float(r.get("ditolak_ha")),
                    "rasio_ditolak": r.get("rasio_ditolak"),
                    "prioritas": r.get("prioritas"),
                    "status_proses": r.get("status_proses"),
                    "status_proses_all": r.get("status_proses_all"),
                    "record_id": r.get("record_id"),
                }
            )
    else:
        for r in sections.get("sk36_2025_110a_riau", {}).get("records", []):
            nama = r.get("Subjek hukum") or r.get("col_1")
            no = parse_numeric(r.get("No") or r.get("col_0"))
            if not nama or no is None:
                continue
            sk36.append(
                {
                    "no": int(no),
                    "nama": nama,
                    "dimohon_ha": parse_numeric(r.get("Dimohon (ha)")),
                    "berproses_ha": parse_numeric(r.get("Berproses 110A (ha)")),
                    "ditolak_ha": parse_numeric(r.get("Ditolak (ha)")),
                    "rasio_ditolak": r.get("Rasio ditolak"),
                    "prioritas": r.get("Prioritas tindak lanjut jika ditolak"),
                    "status_proses": "tidak_diketahui",
                    "record_id": f"sk36-{int(no)}",
                }
            )

    payload = {
        "source": wb_name,
        "updated_note": "Ekspor terstruktur dari workbook penertiban kawasan hutan sawit Riau",
        "normalized": {
            "sebaran_kab_korporasi_kh": {"total": len(kab_kh), "records": kab_kh},
            "gelombang1_27_pt": {"total": len(gelombang1), "records": gelombang1},
            "sk36_2025_110a": {
                "total": len(sk36),
                "records": sk36,
                "pk": "record_id",
                "note": "Composite identity via record_id after DQ fix",
            },
        },
        "sections": sections,
    }
    write_json("penertiban.json", payload)
    print(
        f"    penertiban: kab_kh={len(kab_kh)} gelombang1={len(gelombang1)} sk36={len(sk36)} "
        f"sections={sum(1 for s in sections.values() if s['total'])}"
    )
    return payload


def export_konsesi_gfw_full():
    rows = read_csv("tabulasi_konsesi_sawit_gfw_bbox_riau.csv")
    # Optional geometry centroid lookup from GFW geojson by name/company
    centroids = {}
    gfw_path = ROOT / "tmp" / "spatial" / "gfw_oilpalm_riau.geojson"
    if gfw_path.exists():
        raw = json.loads(gfw_path.read_text(encoding="utf-8"))
        for f in raw.get("features", []):
            props = f.get("properties") or {}
            key = (str(props.get("name") or "").strip().upper(), str(props.get("company") or "").strip().upper())
            lon, lat = geom_centroid_lonlat(f.get("geometry"))
            if lon is not None and lat is not None:
                centroids[key] = {"lon": lon, "lat": lat}

    records = []
    for r in rows:
        name = r.get("name") or r.get("company")
        company = r.get("company")
        key = (str(name or "").strip().upper(), str(company or "").strip().upper())
        key2 = (str(company or "").strip().upper(), str(company or "").strip().upper())
        xy = centroids.get(key) or centroids.get(key2)
        records.append(
            {
                "no": parse_numeric(r.get("no")),
                "company": company,
                "name": name,
                "group": r.get("group_comp"),
                "type": r.get("type"),
                "legal": r.get("po_legalst"),
                "hgu": r.get("po_hgu"),
                "area_hgu_ha": parse_numeric(r.get("po_area_hg")),
                "area_ha": parse_numeric(r.get("area_ha")),
                "source": r.get("source"),
                "gfwid": r.get("gfwid"),
                "catatan": r.get("catatan"),
                "lon": xy.get("lon") if xy else None,
                "lat": xy.get("lat") if xy else None,
            }
        )

    with_xy = sum(1 for r in records if r.get("lon") is not None)
    payload = {
        "source": "tabulasi_konsesi_sawit_gfw_bbox_riau.csv",
        "geometry_source": "tmp/spatial/gfw_oilpalm_riau.geojson (centroid)",
        "total": len(records),
        "with_centroid": with_xy,
        "bbox_note": "Intersect bbox Riau approx; bisa mencakup tepi Sumut/Jambi/Sumbar",
        "records": records,
    }
    write_json("konsesi_gfw_full.json", payload)
    print(f"    konsesi_gfw_full: {len(records)} records, centroid={with_xy}")
    return payload


def export_perusahaan():
    rows = read_csv("daftar_perusahaan_sawit_riau_gabungan.csv")
    records = [
        {
            "no": r.get("no"),
            "nama": r.get("nama_perusahaan"),
            "nama_kanonik": r.get("nama_kanonik") or r.get("nama_perusahaan"),
            "sumber": r.get("sumber"),
            "ada_di_bps": r.get("ada_di_bps"),
            "ada_di_konflik_polda": r.get("ada_di_konflik_polda"),
            "ada_di_gfw": r.get("ada_di_gfw"),
            "ada_di_atlas": r.get("ada_di_atlas"),
            "status_nama": r.get("status_nama"),
            "catatan": r.get("catatan"),
        }
        for r in rows
        if r.get("nama_perusahaan")
    ]
    write_json("perusahaan.json", {"total": len(records), "records": records})


def export_konsesi_atlas():
    gfw = read_csv("tabulasi_konsesi_sawit_gfw_match_bps_riau.csv")
    atlas = read_csv("cocokan_atlas_gabungan_gfw.csv")
    kepmen = read_csv("tabulasi_konsesi_sawit_kepmenhut_36_2025_riau_rapi.csv")
    write_json(
        "konsesi.json",
        {
            "gfw_match_bps": {
                "total": len(gfw),
                "records": [
                    {
                        "nama_bps": r.get("nama_bps_match"),
                        "nama_kanonik": r.get("nama_kanonik"),
                        "company": r.get("company"),
                        "name": r.get("name"),
                        "group": r.get("group_comp"),
                        "type": r.get("type"),
                        "legal": r.get("po_legalst"),
                        "hgu": r.get("po_hgu"),
                        "area_ha": to_float(r.get("area_ha")),
                        "gfwid": r.get("gfwid"),
                    }
                    for r in gfw
                ],
            },
            "atlas_match": {
                "total": len(atlas),
                "records": [
                    {
                        "match_id": r.get("match_id"),
                        "atlas_nama": r.get("atlas_nama"),
                        "atlas_uid": r.get("atlas_uid"),
                        "gfwid": r.get("gfwid"),
                        "tahun": r.get("atlas_tahun"),
                        "tipe": r.get("atlas_tipe"),
                        "status": r.get("status_kecocokan"),
                        "match_method": r.get("match_method"),
                        "match_confidence": r.get("match_confidence"),
                        "nama_lokal": r.get("nama_lokal"),
                        "area_ha": to_float(r.get("area_ha")),
                        "ada_di_bps": r.get("ada_di_bps"),
                        "ada_di_konflik_polda": r.get("ada_di_konflik_polda"),
                        "nama_kanonik": r.get("nama_kanonik"),
                    }
                    for r in atlas
                ],
            },
            "kepmenhut_36_2025": {
                "total": len(kepmen),
                "records": [
                    {
                        "nama": r.get("Nama subjek hukum"),
                        "luas_permohonan_ha": to_float(r.get("Luas permohonan (ha)")),
                        "luas_berproses_ha": to_float(r.get("Luas berproses (ha)")),
                        "luas_ditolak_ha": to_float(r.get("Luas ditolak (ha)")),
                        "status": r.get("Status permohonan"),
                        "kelengkapan": r.get("Kelengkapan data"),
                        "implikasi": r.get("Implikasi"),
                    }
                    for r in kepmen
                ],
            },
        },
    )


def dual_grain_catatan(
    mapped: int,
    lintas: int,
    n_objek: int,
    n_mappable: int,
    n_titik: int,
) -> str:
    return (
        f"Choropleth ADM2; koridor = hull titik objek (default off); densitas centroid (default off); "
        f"coverage {mapped}/{lintas} setelah DQ; skor = indeks liputan+objek+register. "
        f"Bukan batas legal HGU/IUP. objek_titik ({n_titik}) ≠ objek_agrinas ({n_objek}): "
        f"dual grain — titik = proksi spasial (bukan 1:1 dengan registry). "
        f"Metrik benar: objek_mappable ({n_mappable}/{n_objek})."
    )


def count_objek_titik(features: list) -> int:
    return sum(
        1
        for f in features
        if (f.get("properties") or {}).get("layer") == "objek_titik"
    )


def count_objek_mappable(objek: list[dict]) -> int:
    return sum(
        1
        for r in objek
        if str(r.get("mappable") or "").strip().lower() in {"ya", "true", "1", "yes"}
    )


def export_meta(counts: dict, catatan: str | None = None):
    write_json(
        "meta.json",
        {
            "brand": "Atlas Konflik Sawit Riau",
            "subtitle": "Pemetaan spasial konflik, objek Agrinas–Satgas, dan celahan perizinan",
            "updated_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
            "bbox_riau": list(RIAU_BBOX),
            "center": [101.45, 0.55],
            "zoom": 8,
            "counts": counts,
            "layers": [
                {"id": "choropleth", "label": "Choropleth kab/kota", "default": True},
                {"id": "koridor", "label": "Koridor proksi", "default": False},
                {"id": "densitas_kasus", "label": "Densitas kasus", "default": False},
                {"id": "objek_titik", "label": "Titik objek Agrinas", "default": True},
                {"id": "gfw_konsesi", "label": "Konsesi GFW", "default": False},
            ],
            "views": ["peta", "analisis", "cerita", "data"],
            "methodology": {
                "skor_type": "indeks_liputan_objek_register",
                "disclaimer": (
                    "Indeks liputan+objek+register terbuka — bukan vonis operasional. "
                    f"Setelah DQ: {counts.get('entri_terpetakan', '–')}/"
                    f"{counts.get('entri_tidak_terpetakan', '–')} entri terpetakan/tidak. "
                    "Kalibrasi ulang dengan rekap LP/SPKT 36 bulan resmi."
                ),
                "blend_default": "70% OSINT + 30% register",
                "kalibrasi": "Menunggu rekap LP/SPKT 36 bulan resmi",
                "dq_note": (
                    "DQ plan applied: noise kasus dropped; tanpa_lp flag; kab_primary pada objek; "
                    "match_id cocokan; sk36 record_id; company alias."
                ),
                "dual_grain": (
                    "objek_agrinas = entity registry; objek_titik = proksi spasial campuran; "
                    "objek_mappable = subset registry yang layak diplot. Jangan bandingkan titik vs registry 1:1."
                ),
            },
            "sumber": [
                "TABEL_KONFLIK_AGRARIA_SAWIT_RIAU",
                "Master_List_Objek_Agrinas_Satgas_Riau",
                "Ranking_Potensi_Konflik_Per_Polres",
                "cluster_kabkota_agrinas",
                "idn_adm2_simplified (choropleth)",
                "gfw_oilpalm_riau (TopoJSON overlay, lazy)",
                "Tabulasi_Kepmenhut_36_2025",
                "Tabulasi_Penertiban_Kawasan_Hutan_Sawit_Riau",
                "tabulasi_konsesi_sawit_gfw_bbox_riau (287)",
                "Nusantara Atlas / GFW (cocokan)",
            ],
            "update_command": (
                "python website/scripts/apply_dq_fixes.py && python website/scripts/export_web_data.py"
            ),
            "catatan": catatan
            or dual_grain_catatan(
                counts.get("entri_terpetakan", 0),
                counts.get("entri_tidak_terpetakan", 0),
                counts.get("objek_agrinas", 0),
                counts.get("objek_mappable", 0),
                counts.get("objek_titik", 0),
            ),
        },
    )


def main():
    print(f"ROOT = {ROOT}")
    print(f"OUT  = {OUT}")
    kab_records = export_kab_kota()
    polres = export_polres()
    objek = export_objek()
    kasus = export_kasus()
    patch_polres_coverage(kasus, polres)
    attach_kasus_counts(kab_records, kasus)
    write_json("kab_kota.json", {"updated": True, "records": kab_records})
    adm = export_adm2_choropleth(kab_records)
    # rewrite kab after shape_name/n_kasus filled
    write_json("kab_kota.json", {"updated": True, "records": kab_records})
    export_perusahaan()
    export_konsesi_atlas()
    export_penertiban()
    gfw_full = export_konsesi_gfw_full()
    geo = export_spatial_layers(kab_records)
    gfw = export_gfw_overlay()
    export_analytics(polres, objek, kasus)
    lintas = sum(
        1
        for r in kasus
        if "lintas" in str(r.get("polres") or "").lower()
        or "lintas" in str(r.get("kab_kota") or "").lower()
    )
    n_mappable = count_objek_mappable(objek)
    n_titik = count_objek_titik(geo["features"])
    counts = {
        "kab_kota": len(kab_records),
        "polres": len(polres),
        "objek_agrinas": len(objek),
        "objek_mappable": n_mappable,
        "objek_titik": n_titik,
        "kasus_konflik": len(kasus),
        "choropleth": len(adm),
        "fitur_spasial": len(geo["features"]),
        "gfw_konsesi": len(gfw),
        "gfw_bbox_full": gfw_full.get("total", 0),
        "entri_terpetakan": len(kasus) - lintas,
        "entri_tidak_terpetakan": lintas,
    }
    export_meta(
        counts,
        catatan=dual_grain_catatan(
            counts["entri_terpetakan"],
            counts["entri_tidak_terpetakan"],
            counts["objek_agrinas"],
            counts["objek_mappable"],
            counts["objek_titik"],
        ),
    )
    # DQ gate + report
    try:
        from validate_web_data import main as validate_main

        rc = validate_main()
        if rc != 0:
            raise SystemExit(f"validate_web_data failed with code {rc}")
    except ImportError:
        print("WARN: validate_web_data not importable")
    try:
        from write_dq_report import main as dq_report_main

        dq_report_main()
    except ImportError:
        print("WARN: write_dq_report not available yet")
    print("Selesai. Refresh website untuk melihat data terbaru.")


if __name__ == "__main__":
    main()
