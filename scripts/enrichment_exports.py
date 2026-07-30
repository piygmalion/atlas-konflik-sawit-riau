#!/usr/bin/env python3
"""Enrichment exports Fase 1–3 (additive gold blobs / field patches)."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

HERE = Path(__file__).resolve()
sys.path.insert(0, str(HERE.parent))

from company_normalize import load_alias_table, resolve_canonical  # noqa: E402

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl diperlukan: pip install openpyxl") from exc


def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    return s if s and s.lower() not in {"none", "null", "-"} else None


def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s and s not in {".", "-", "-."} else None
    except ValueError:
        return None


def _slug(*parts: str) -> str:
    raw = "-".join(str(p or "").strip() for p in parts if p)
    s = re.sub(r"[^a-zA-Z0-9]+", "-", raw).strip("-").lower()
    return s or "unknown"


def _norm_kab(name: str | None) -> str:
    s = re.sub(r"\s+", " ", str(name or "").strip().lower())
    s = re.sub(r"^(kabupaten|kota|kab\.?)\s+", "", s)
    return s


def export_perusahaan_alias(root: Path, out: Path) -> dict:
    path = root / "dim_perusahaan_alias.csv"
    records = []
    if path.exists():
        import csv

        with path.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                mentah = _clean(row.get("nama_mentah"))
                kanon = _clean(row.get("nama_kanonik"))
                if not mentah or not kanon:
                    continue
                records.append(
                    {
                        "nama_mentah": mentah,
                        "nama_kanonik": kanon,
                        "sumber": _clean(row.get("sumber")),
                        "confidence": _clean(row.get("confidence")),
                    }
                )
    # Dedup PK case-insensitive keep first
    seen = set()
    uniq = []
    for r in records:
        key = r["nama_mentah"].lower()
        if key in seen:
            continue
        seen.add(key)
        uniq.append(r)
    payload = {"total": len(uniq), "records": uniq}
    (out / "perusahaan_alias.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  wrote perusahaan_alias.json ({len(uniq)} aliases)")
    return payload


def build_atlas_full(root: Path) -> dict:
    import csv

    path = root / "tabulasi_konsesi_sawit_nusantara_atlas_riau.csv"
    records = []
    if not path.exists():
        return {"total": 0, "records": []}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            nama = _clean(row.get("nama_perusahaan"))
            if not nama:
                continue
            uid = _clean(row.get("uid"))
            no = _clean(row.get("no"))
            kab = _clean(row.get("kabupaten"))
            atlas_id = uid or (f"ATLAS-{no}" if no is not None else None)
            if not atlas_id:
                atlas_id = "ATLAS-" + hashlib.sha1(
                    f"{nama}|{kab}".encode("utf-8")
                ).hexdigest()[:12]
            records.append(
                {
                    "atlas_id": atlas_id,
                    "no": no,
                    "nama_perusahaan": nama,
                    "nama_kanonik": _clean(row.get("nama_kanonik")) or nama,
                    "grup": _clean(row.get("grup")),
                    "parent_group": _clean(row.get("parent_group")),
                    "subgroup": _clean(row.get("subgroup")),
                    "kabupaten": kab,
                    "luas_ha": _to_float(row.get("luas_ha")),
                    "luas_gambut_ha": _to_float(row.get("luas_gambut_ha")),
                    "hutan_tersisa_ha": _to_float(row.get("hutan_tersisa_ha")),
                    "tipe_konsesi": _clean(row.get("tipe_konsesi")),
                    "komoditas": _clean(row.get("komoditas")),
                    "nomor_izin": _clean(row.get("nomor_izin")),
                    "tipe_izin": _clean(row.get("tipe_izin")),
                    "status_izin": _clean(row.get("status_izin")),
                    "luas_izin_ha": _to_float(row.get("luas_izin_ha")),
                    "ispo": _clean(row.get("ispo")),
                    "rspo": _clean(row.get("rspo")),
                    "uid": uid,
                    "cid": _clean(row.get("cid")),
                }
            )
    return {"total": len(records), "records": records}


def atlas_full_analytics(atlas_full: dict) -> dict:
    from collections import Counter

    rows = atlas_full.get("records") or []
    by_kab = Counter(_clean(r.get("kabupaten")) or "Tidak diketahui" for r in rows)
    by_grup = Counter(_clean(r.get("grup")) or "Tanpa grup" for r in rows)
    links = []
    for kab, n in by_kab.most_common():
        links.append({"source": "Nusantara Atlas (full)", "target": kab, "value": n})
    top_grup = by_grup.most_common(12)
    for grup, n in top_grup:
        links.append({"source": "Grup Atlas", "target": grup, "value": n})
    return {
        "links": links,
        "by_kabupaten": [{"label": k, "value": v} for k, v in by_kab.most_common()],
        "by_grup": [{"label": k, "value": v} for k, v in top_grup],
        "total": len(rows),
        "grain": "atlas_full",
    }


def enrich_kab_verifikasi(root: Path, kab_records: list[dict]) -> int:
    """Additive fields from verifikasi + perkiraan sebaran workbooks."""
    path = root / "tabulasi_verifikasi_lanjutan_sebaran_riau.xlsx"
    perkiraan = root / "tabulasi_perkiraan_lokasi_sebaran_riau.xlsx"
    ver_map: dict[str, dict] = {}
    if path.exists():
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["01_Verifikasi Kabupaten"] if "01_Verifikasi Kabupaten" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if rows:
            headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
                kab = d.get("Kabupaten/Kota") or d.get(headers[0])
                if not kab:
                    continue
                ver_map[_norm_kab(kab)] = {
                    "verifikasi_status": d.get("Status verifikasi"),
                    "kepercayaan_sebaran": d.get("Kepercayaan"),
                    "rank_gfw": d.get("Rank GFW area"),
                    "rank_sebaran": d.get("Rank sebaran peta"),
                }
    if perkiraan.exists() and not ver_map:
        wb = openpyxl.load_workbook(perkiraan, read_only=True, data_only=True)
        sheet = "Ringkasan Kabupaten" if "Ringkasan Kabupaten" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if rows:
            headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
                kab = d.get("Kabupaten/Kota")
                if not kab:
                    continue
                ver_map[_norm_kab(kab)] = {
                    "verifikasi_status": d.get("Intensitas Sebaran"),
                    "kepercayaan_sebaran": d.get("Perkiraan Kepadatan Relatif"),
                    "rank_gfw": None,
                    "rank_sebaran": d.get("No"),
                }
    n = 0
    for rec in kab_records:
        key = _norm_kab(rec.get("kab_kota"))
        hit = ver_map.get(key)
        if not hit:
            for vk, vv in ver_map.items():
                if key in vk or vk in key:
                    hit = vv
                    break
        if not hit:
            rec.setdefault("verifikasi_status", None)
            rec.setdefault("kepercayaan_sebaran", None)
            rec.setdefault("rank_gfw", None)
            rec.setdefault("rank_sebaran", None)
            continue
        rec["verifikasi_status"] = hit.get("verifikasi_status")
        rec["kepercayaan_sebaran"] = hit.get("kepercayaan_sebaran")
        rec["rank_gfw"] = hit.get("rank_gfw")
        rec["rank_sebaran"] = hit.get("rank_sebaran")
        n += 1
    print(f"    kab verifikasi enrich: {n}/{len(kab_records)}")
    return n


def hotspot_verifikasi_features(root: Path) -> list[dict]:
    """Inject only georef hotspots with verification status (no REF placeholders)."""
    path = root / "tabulasi_verifikasi_lanjutan_sebaran_riau.xlsx"
    features = []
    if not path.exists():
        return features
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "02_Hotspot Georef" not in wb.sheetnames:
        wb.close()
        return features
    ws = wb["02_Hotspot Georef"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return features
    headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[0])]
    for row in rows[1:]:
        d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
        status = str(d.get("Status") or "")
        if not re.search(r"terkonfirmasi|terverifikasi", status, re.I):
            continue
        lon = _to_float(d.get("Lon"))
        lat = _to_float(d.get("Lat"))
        if lon is None or lat is None:
            continue
        hid = d.get("ID") or f"HS-{len(features)+1}"
        features.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": {
                    "id": hid,
                    "nama": f"Hotspot {hid}",
                    "kab_kota": d.get("Kabupaten (PIP)"),
                    "kecamatan": d.get("Kecamatan Nominatim"),
                    "pixels": d.get("Piksel"),
                    "status": status,
                    "osint": d.get("Cocokan OSINT"),
                    "layer": "hotspot_verifikasi",
                    "tipe": "hotspot_georef",
                },
            }
        )
    print(f"    hotspot_verifikasi features: {len(features)}")
    return features


# ─── Fase 2 ──────────────────────────────────────────────────────────────────

SHEET_KAB_IZIN = {
    "kampar": "Kabupaten Kampar",
    "ROHUL": "Kabupaten Rokan Hulu",
    "ROHIL": "Kabupaten Rokan Hilir",
    "DUMAI": "Kota Dumai",
    "MERANTI": "Kabupaten Kepulauan Meranti",
    "BENGKLAIS": "Kabupaten Bengkalis",
    "SIAK": "Kabupaten Siak",
    "PELALAWAN": "Kabupaten Pelalawan",
    "INHU": "Kabupaten Indragiri Hulu",
    "INHIL": "Kabupaten Indragiri Hilir",
    "KUANSING": "Kabupaten Kuantan Singingi",
    "PEKANBARU": "Kota Pekanbaru",
}


def _izin_to_num(v) -> float:
    x = _to_float(v)
    return float(x) if x is not None else 0.0


def _is_address_or_noise(text: str) -> bool:
    t = text.strip()
    if not t:
        return True
    u = t.upper()
    if u.startswith(("JL", "JALAN", "JL.", "GEDUNG", "KOMP", "KOMPLEK", "BLOK", "MENARA")):
        return True
    if u.startswith(("-", "–", "—")):
        return True
    if re.match(r"^\(?PIR", u):
        return True
    if not re.search(r"\b(PT|CV|PTPN|KUD|KOPERASI)\b", u) and (
        "PEKANBARU" in u or "JAKARTA" in u or u == "RIAU" or "NO." in u
    ):
        return True
    return False


def _is_company_start(no_val, name_val) -> bool:
    if name_val is None:
        return False
    name = str(name_val).strip()
    if not name or _is_address_or_noise(name):
        return False
    if isinstance(no_val, (int, float)) and not isinstance(no_val, bool):
        return True
    if isinstance(no_val, str) and re.fullmatch(r"\d+", no_val.strip()):
        return True
    u = name.upper()
    if re.match(r"^(PT|CV|PTPN|KUD|KOPERASI)\b", u):
        return True
    return False


def _clean_company_name(raw: str) -> str:
    s = re.sub(r"\s+", " ", str(raw)).strip(" /")
    s = re.sub(r"\s*\+\s*PMKS.*$", "", s, flags=re.I)
    s = re.sub(r"^PT\.?\s*", "PT ", s, flags=re.I)
    s = re.sub(r"^CV\.?\s*", "CV ", s, flags=re.I)
    s = re.sub(r"^PTPN\.?\s*", "PTPN ", s, flags=re.I)
    s = re.sub(r"^KUD\.?\s*", "KUD ", s, flags=re.I)
    return s.upper().strip()


def _parse_izin_kab_sheet(ws, kabupaten: str) -> list[dict]:
    rows = []
    current = None
    name_parts: list[str] = []

    def flush():
        nonlocal current, name_parts
        if not current:
            return
        full = re.sub(r"\s+", " ", " ".join(name_parts).strip())
        parts = full.split(" ")
        kept = []
        for p in parts:
            if _is_address_or_noise(p) and kept:
                break
            kept.append(p)
        nama = _clean_company_name(" ".join(kept))
        if not nama or nama in {"PT", "CV", "JUMLAH", "TOTAL"} or nama.upper().startswith("JUMLAH"):
            current = None
            name_parts = []
            return
        current["nama_mentah"] = nama
        current["nama_asli"] = " ".join(name_parts)
        rows.append(current)
        current = None
        name_parts = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c3, c4, c5, c6, c7, c8 = [ws.cell(r, i).value for i in range(3, 9)]
        if isinstance(a, str) and a.strip().upper() in {
            "NO",
            "REKAPITULASI PERIZINAN PERKEBUNAN",
            "JUMLAH",
            "TOTAL",
        }:
            if str(a).strip().upper() in {"JUMLAH", "TOTAL"}:
                flush()
            continue
        if isinstance(b, str) and b.strip().upper() == "NAMA PERUSAHAAN":
            continue
        if _is_company_start(a, b):
            flush()
            current = {
                "kabupaten": kabupaten,
                "izin_lokasi_ha": 0.0,
                "iup_ha": 0.0,
                "pelepasan_kh_ha": 0.0,
                "hgu_ha": 0.0,
                "pks_unit": 0.0,
                "pks_kapasitas_tbs": 0.0,
            }
            name_parts = [str(b).strip()]
            current["izin_lokasi_ha"] += _izin_to_num(c3)
            current["iup_ha"] += _izin_to_num(c4)
            current["pelepasan_kh_ha"] += _izin_to_num(c5)
            current["hgu_ha"] += _izin_to_num(c6)
            current["pks_unit"] += _izin_to_num(c7)
            current["pks_kapasitas_tbs"] += _izin_to_num(c8)
            continue
        if current is None:
            continue
        if b is not None:
            bt = str(b).strip()
            if bt and not _is_address_or_noise(bt):
                if re.match(r"^[A-Za-z].*", bt) and not re.match(r"^\d", bt):
                    if len(bt) < 40 and "JL" not in bt.upper():
                        name_parts.append(bt)
        current["izin_lokasi_ha"] += _izin_to_num(c3)
        current["iup_ha"] += _izin_to_num(c4)
        current["pelepasan_kh_ha"] += _izin_to_num(c5)
        current["hgu_ha"] += _izin_to_num(c6)
        current["pks_unit"] += _izin_to_num(c7)
        current["pks_kapasitas_tbs"] += _izin_to_num(c8)
    flush()
    return rows


def export_izin_2017(root: Path, out: Path) -> dict:
    path = root / "REKAPITULASI PERIZINAN PERKEBUNAN 2017 ok.xlsx"
    alias_map = load_alias_table(root / "dim_perusahaan_alias.csv")
    records = []
    if path.exists():
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        for sheet_name, kab in SHEET_KAB_IZIN.items():
            if sheet_name not in wb.sheetnames:
                continue
            for row in _parse_izin_kab_sheet(wb[sheet_name], kab):
                nama = row["nama_mentah"]
                kab_id = _slug(kab.replace("Kabupaten ", "").replace("Kota ", ""))
                record_id = "IZ17-" + hashlib.sha1(
                    f"{kab_id}|{nama}".encode("utf-8")
                ).hexdigest()[:12]
                records.append(
                    {
                        "record_id": record_id,
                        "kab_id": kab_id,
                        "kab_kota": kab,
                        "nama_mentah": nama,
                        "nama_kanonik": resolve_canonical(nama, alias_map),
                        "izin_lokasi_ha": row["izin_lokasi_ha"] or None,
                        "iup_ha": row["iup_ha"] or None,
                        "pelepasan_kh_ha": row["pelepasan_kh_ha"] or None,
                        "hgu_ha": row["hgu_ha"] or None,
                        "pks_unit": row["pks_unit"] or None,
                        "pks_kapasitas_tbs": row["pks_kapasitas_tbs"] or None,
                        "vintage": 2017,
                    }
                )
        wb.close()
    # unique PK
    seen = set()
    uniq = []
    for r in records:
        if r["record_id"] in seen:
            continue
        seen.add(r["record_id"])
        uniq.append(r)
    payload = {
        "total": len(uniq),
        "vintage": 2017,
        "disclaimer": (
            "Rekap perizinan perkebunan vintage 2017 — bukan status izin terkini. "
            "Grain: perusahaan × kabupaten."
        ),
        "records": uniq,
    }
    (out / "izin_2017.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  wrote izin_2017.json ({len(uniq)} records)")
    return payload


def export_desa_lock(root: Path, out: Path) -> dict:
    records = []
    summary = root / "tmp" / "spatial" / "desa_lock" / "summary_kunci_desa_final.json"
    xlsx = root / "tabulasi_kunci_desa_V02_V11_V12_V20.xlsx"
    if summary.exists():
        raw = json.loads(summary.read_text(encoding="utf-8"))
        for r in raw:
            records.append(
                {
                    "id": r.get("id"),
                    "kabupaten": r.get("kab"),
                    "kecamatan": r.get("kec"),
                    "desa": r.get("desa"),
                    "desa_utama": r.get("desa_utama"),
                    "lon": r.get("lon"),
                    "lat": r.get("lat"),
                    "dist_km": r.get("dist_km"),
                    "kepercayaan": r.get("conf"),
                    "tetangga": r.get("tetangga"),
                    "metode": r.get("metode"),
                    "sent_scene": r.get("sent_scene"),
                    "sent_date": r.get("sent_date"),
                    "cloud_pct": r.get("cloud"),
                    "sent_vis": r.get("sent_vis"),
                    "bukti": r.get("bukti"),
                }
            )
    elif xlsx.exists():
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb["Kunci Desa Final"] if "Kunci Desa Final" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if rows:
            headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
                rid = d.get("ID")
                if not rid:
                    continue
                records.append(
                    {
                        "id": rid,
                        "kabupaten": d.get("Kabupaten"),
                        "kecamatan": d.get("Kecamatan"),
                        "desa": d.get("Desa kunci"),
                        "desa_utama": d.get("Desa utama"),
                        "lon": _to_float(d.get("Lon")),
                        "lat": _to_float(d.get("Lat")),
                        "dist_km": _to_float(d.get("Jarak ke pusat desa (km)")),
                        "kepercayaan": d.get("Kepercayaan"),
                        "tetangga": d.get("Tetangga desa (km)"),
                        "metode": d.get("Metode"),
                        "sent_scene": d.get("Sentinel scene"),
                        "sent_date": str(d.get("Tanggal citra") or "") or None,
                        "cloud_pct": _to_float(d.get("Cloud %")),
                        "sent_vis": d.get("Interpretasi visual Sentinel"),
                        "bukti": None,
                    }
                )
    payload = {"total": len(records), "records": records}
    (out / "desa_lock.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  wrote desa_lock.json ({len(records)} records)")
    return payload


def _norm_co(s: str | None) -> str:
    t = re.sub(r"[^a-z0-9]+", " ", str(s or "").lower())
    t = re.sub(r"\b(pt|cv|tbk|persero)\b", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def apply_fase2_objek_flags(root: Path, objek: list[dict]) -> int:
    """Additive fase2_gap / mitra_eval on objek_agrinas (null-ok)."""
    path = root / "Master_List_Fase2_Agrinas_Satgas_Riau.xlsx"
    for o in objek:
        o.setdefault("fase2_gap", None)
        o.setdefault("mitra_eval", None)
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    gap_notes: dict[str, str] = {}
    mitra_notes: dict[str, str] = {}

    if "07_Gap_Matching_Sumber" in wb.sheetnames:
        ws = wb["07_Gap_Matching_Sumber"]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row containing 'Gap' or 'Sumber'
        header_i = 0
        for i, row in enumerate(rows[:8]):
            blob = " ".join(str(c or "") for c in row).lower()
            if "gap" in blob or "sumber" in blob or "entitas" in blob:
                header_i = i
                break
        headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[header_i])]
        for row in rows[header_i + 1 :]:
            d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
            vals = [v for v in d.values() if v]
            if len(vals) < 2:
                continue
            nama = vals[0]
            note = " | ".join(str(v) for v in vals[1:4])
            gap_notes[_norm_co(nama)] = note

    if "03_Mitra_KSO_Evaluasi" in wb.sheetnames:
        ws = wb["03_Mitra_KSO_Evaluasi"]
        rows = list(ws.iter_rows(values_only=True))
        header_i = 3 if len(rows) > 4 else 0
        for i, row in enumerate(rows[:6]):
            blob = " ".join(str(c or "") for c in row).lower()
            if "mitra" in blob and ("utama" in blob or "tanggal" in blob):
                header_i = i
                break
        headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[header_i])]
        for row in rows[header_i + 1 :]:
            d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
            mitra = d.get("Mitra utama") or next(
                (v for k, v in d.items() if v and "mitra" in k.lower()), None
            )
            if not mitra:
                continue
            label = d.get("Label") or d.get("Kab (indikasi)") or "evaluasi"
            mitra_notes[_norm_co(mitra)] = str(label)

    if "02_Master_List_Inti" in wb.sheetnames:
        ws = wb["02_Master_List_Inti"]
        rows = list(ws.iter_rows(values_only=True))
        header_i = 3 if len(rows) > 4 else 0
        for i, row in enumerate(rows[:6]):
            blob = " ".join(str(c or "") for c in row).lower()
            if "nama" in blob or "entitas" in blob:
                header_i = i
                break
        headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[header_i])]
        for row in rows[header_i + 1 :]:
            d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
            nama = d.get("Nama") or d.get("Nama entitas") or next(
                (v for k, v in d.items() if v and "nama" in k.lower()), None
            )
            label = d.get("Label") or d.get("Kelas") or d.get("Status")
            if nama and label:
                gap_notes.setdefault(_norm_co(nama), str(label))

    wb.close()
    n = 0
    for o in objek:
        on = _norm_co(o.get("nama"))
        if not on:
            continue
        gap = gap_notes.get(on)
        mitra = mitra_notes.get(on)
        if not gap:
            for k, v in gap_notes.items():
                if k and on and (k in on or on in k):
                    gap = v
                    break
        if not mitra:
            for k, v in mitra_notes.items():
                if k and on and (k in on or on in k):
                    mitra = v
                    break
        if gap:
            o["fase2_gap"] = gap
            n += 1
        if mitra:
            o["mitra_eval"] = mitra
            n += 1
    print(f"    objek fase2 flags applied touches={n}")
    return n


def patch_izin_flags(perusahaan_path: Path, kab_path: Path, izin: dict) -> None:
    """Additive ada_izin_2017 on perusahaan + n_izin_2017 on kab."""
    izin_kanon = {_norm_co(r.get("nama_kanonik") or r.get("nama_mentah")) for r in izin.get("records") or []}
    by_kab = defaultdict(int)
    for r in izin.get("records") or []:
        by_kab[_norm_kab(r.get("kab_kota"))] += 1

    if perusahaan_path.exists():
        data = json.loads(perusahaan_path.read_text(encoding="utf-8"))
        for r in data.get("records") or []:
            key = _norm_co(r.get("nama_kanonik") or r.get("nama"))
            r["ada_izin_2017"] = bool(key and key in izin_kanon) or any(
                key and (key in k or k in key) for k in izin_kanon if k
            )
        perusahaan_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    if kab_path.exists():
        data = json.loads(kab_path.read_text(encoding="utf-8"))
        for r in data.get("records") or []:
            key = _norm_kab(r.get("kab_kota"))
            n = by_kab.get(key, 0)
            if not n:
                for bk, bv in by_kab.items():
                    if key in bk or bk in key:
                        n = bv
                        break
            r["n_izin_2017"] = n
        kab_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def export_rantai_agrinas(root: Path, out: Path) -> dict:
    """Fase 3 mart: baseline rantai Satgas → Agrinas (+ grup Atlas ringkas)."""
    path = root / "Baseline_Publik_Rantai_Satgas_Agrinas.xlsx"
    stages = []
    timeline = []
    klaim = []
    if path.exists():
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        def sheet_table(name: str) -> list[dict]:
            if name not in wb.sheetnames:
                return []
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 4:
                return []
            # title rows 1-2, header ~4
            header_i = 3 if len(rows) > 3 else 0
            for i, row in enumerate(rows[:6]):
                blob = " ".join(str(c or "") for c in row).lower()
                if any(tok in blob for tok in ("tahap", "tahun", "klaim", "item", "uraian", "no")):
                    header_i = i
            headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[header_i])]
            out_rows = []
            for row in rows[header_i + 1 :]:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                d = {headers[i]: _clean(row[i] if i < len(row) else None) for i in range(len(headers))}
                if any(d.values()):
                    out_rows.append(d)
            return out_rows

        stages = sheet_table("02_Tahap_Penyerahan")
        timeline = sheet_table("03_Timeline")
        klaim = sheet_table("04_Klaim_Riau")
        wb.close()

    grup_path = root / "tabulasi_grup_konsesi_sawit_nusantara_atlas_riau.csv"
    grup = []
    if grup_path.exists():
        import csv

        with grup_path.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                grup.append({k: _clean(v) for k, v in row.items()})

    payload = {
        "total_stages": len(stages),
        "stages": stages,
        "timeline": timeline,
        "klaim_riau": klaim,
        "atlas_grup": {"total": len(grup), "records": grup[:50]},
        "disclaimer": "Baseline publik rantai Satgas→Agrinas — bukan daftar BA resmi.",
    }
    (out / "rantai_agrinas.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  wrote rantai_agrinas.json (stages={len(stages)} grup={len(grup)})")
    return payload


def write_silver_staging(out: Path, payloads: dict[str, list[dict] | dict]) -> None:
    """Local silver mirror used by materialize_serving (Fase 3)."""
    silver = out / "silver"
    silver.mkdir(parents=True, exist_ok=True)
    for name, payload in payloads.items():
        path = silver / f"{name}.json"
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  silver staging: {len(payloads)} tables -> {silver}")
